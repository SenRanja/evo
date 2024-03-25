# encoding=utf-8
import json
import os
import time

from openpyxl import load_workbook

from conf.errors import error_log_msg, success_msg, error_license_expired_msg, error_excel_import_user_msg, \
    error_no_active_user_login_msg
from django.contrib import auth
from django.db.models import Q
from django.http import JsonResponse as JR
from django.contrib.auth.models import Group
from guardian.shortcuts import assign_perm
from lib.handle_license_data import check_license_data_valid
from exam_system.settings import EXCEL_IMPORT_QUESTION_ROOT
from user.models import User
from module.views import generate_user_info

# 【用户 User】
# 注册用户
def register(request):
    data = json.loads(request.body)

    name = data["name"]
    pwd = data['password']
    stu_id = data['stu_id']
    tel = data['tel']
    id_card = data['id_card']
    email = data['email']
    role = data['role']

    # 注册超级管理员
    if data.get("super") == True:
        user_status = "admin"
        user = User.objects.create_superuser(password=pwd, name=name, stu_id=stu_id, tel=tel, id_card=id_card, email=email,role=role)
    else:
        # 注册教师
        if data.get("role") == "tea":
            user_status = "teacher"
            user = User.objects.create_user(password=pwd,name=name,stu_id=stu_id,tel=tel,id_card=id_card,email=email,role=role)
        # 注册学生用户
        elif data.get("role") == "stu":
            user_status = "student"
            user = User.objects.create_user(password=pwd,name=name,stu_id=stu_id,tel=tel,id_card=id_card,email=email,role=role)
        else:
            return JR(error_log_msg("未填写用户是学生还是教师"))
    return JR(success_msg("created new user: {name}".format(name=name)))


# 用户登录
def login(request):
    # 若用户被禁用(即is_active==0)，可以登录但是session无效
    data = json.loads(request.body)

    user = auth.authenticate(username=data['login_name'], password=data['password'])
    if user == None:
        res = JR(error_log_msg("login failed"))
        res.delete_cookie('session_id')
        return res
    if user.is_active==False:
        res = JR(error_no_active_user_login_msg(request, "用户被禁用"))
        res.delete_cookie('session_id')
        return res
    else:
        auth.login(request, user)

        if user.is_superuser == False:
            # 每次触发非管理员用户的登录，均验证系统license是否过期。若过期则不处理登录的正常业务。
            license_valid_bool, license_valid_reason = check_license_data_valid()
            if license_valid_bool == False:
                res = JR(error_license_expired_msg(license_valid_reason))
                res.delete_cookie('session_id')
                return res

        user_msg = {
            'id': user.id,
            'name': user.name,
            'stu_id': user.stu_id,
            'tel': user.tel,
            'id_card': user.id_card,
            'email': user.email,
            'create_time': user.create_time,
            'role': user.role,
        }
        ret_dict = success_msg("login successfully")
        ret_dict.update(user_msg)
        ret = JR(ret_dict)

        return ret


# 用户登出
def logout(request):
    name = str(request.user)
    auth.logout(request)
    return JR(success_msg(name + " logout"))


# 用户登录验证
def login_check(request):
    "利用中间件 AuthCookie，如果响应401说明未登录，如果响应200说明登录成功，通过验证"
    user = request.user
    try:
        user_msg = {
            'id': user.id,
            'name': user.name,
            'stu_id': user.stu_id,
            'tel': user.tel,
            'id_card': user.id_card,
            'email': user.email,
            'create_time': user.create_time,
            'role': user.role,
        }
        print("用户自身权限")
        perms = user.get_user_permissions()
        print(perms)
        print("用户带组自身权限")
        perms = user.get_all_permissions()
        print(perms)
        print("用户对特定应用权限")
        # view_exammanage  add_exammanage  change_exammanage  delete_exammanage
        perms = user.has_perm('exam_manage.view_exammanage')
        print(perms)
        print("用户对特定应用的对象级权限")
        # print("给予权限")
        # assign_perm('exam_manage.view_exammanage', user)
        return JR(user_msg)
    except:
        name = 'anonymous'
        return JR(success_msg(name))

# 改密
def updatepassword(request):
    data = json.loads(request.body)
    user = request.user
    stu_id = user.stu_id
    if user.check_password(data['oldpassword']):
        if data['password']==data['repassword']:
            user.set_password(data['password'])
            user.save()
            return JR({'success': True, 'message': '密码已更新'})
        else:
            return JR({'success': False, 'message': '新密码和确认密码不匹配'})
    else:
        return JR({'success': False, 'message': '旧密码错误'})


# 批量导入用户数据
def import_user_from_excel(request):
    try:
        excel_file = request.FILES['excel_file']

        user = request.user
        id = user.id
        user_name = user.name
        # print(excel_file)  # 打印文件名

        # 将文件保存到某位置
        current_time = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
        filename = "用户导入_{current_time}_{id}_{name}.xlsx".format(current_time=current_time,id=id,name=user_name)
        file_path = os.path.join(EXCEL_IMPORT_QUESTION_ROOT, filename)
        with open(file_path, 'wb') as f:
            for chunk in excel_file.chunks():
                f.write(chunk)

        # excel文件中内容数据入库
        wb = load_workbook(excel_file)
        handle_sheet = wb.worksheets[0]
        # 从第二行开始读取
        for row in handle_sheet.iter_rows(min_row=2, values_only=True):
            try:
                new_user_obj = User.objects.create(
                    name=row[0],
                    stu_id=row[1],
                    tel=row[3],
                    id_card=row[4],
                    email=row[5],
                    role=row[6],
                )
                # 设置密码不能在create方法直接指定password，否则密码是未加密的明文存储，必须如下单独set_password才可
                new_user_obj.set_password(row[7])
                new_user_obj.save()

                # 如果 组 不存在，就建立
                if Group.objects.filter(name=row[2]).exists():
                    group = Group.objects.get(name=row[2])
                    new_user_obj.groups.add(group)
                else:
                    group = Group.objects.create(name=row[2])
                    new_user_obj.groups.add(group)
            except Exception as err:
                if "Duplicate entry" in str(err):
                    pass

        # 手动关闭文件句柄
        excel_file.close()

        return JR(success_msg('import_user_from_excel successfully'))
    except Exception as err:
        return JR(error_excel_import_user_msg(request, 'import_user_from_excel failed: '+str(err)))


# 用户更改密码
def change_passwd(request):
    '''非admin权限用户的改密接口'''
    data = json.loads(request.body)
    user = request.user
    old_password = data["old_password"]
    new_password = data["new_password"]
    if user.check_password(old_password)==True:
        user.set_password(new_password)
        # save()后默认logout，
        user.save()
        # 因此不需要手动logout
        return JR(success_msg("change pwd successfully"))
    else:
        return JR(success_msg("old_password wrong"))

# 【组 group】
# 用户可以自行添加组，但是仅admin权限可创建、删除、改变组
def group_query(request):
    user = request.user
    print(user.groups.all())
    return JR(success_msg('okk2'))

def group_add(request):
    data = json.loads(request.body)

    grp_name = data['grp_name']
    try:
        custom_group = Group.objects.create(name=grp_name,)
        return JR(success_msg("oko"))
    except Exception as err:
        return JR(success_msg(str(err)))

def group_add_user(request):
    data = json.loads(request.body)

    grp_name = data['grp_name']
    group, created = Group.objects.get_or_create(name=grp_name)

    user = request.user
    print(user.groups.all())
    user.groups.add(group)
    print(user.groups.all())
    return JR(success_msg('k'))

def group_add_user_by_sid(request, group_id:int, sid:str):
    # sid 是学生stu_id 或 身份证号
    user = request.user
    grp = Group.objects.get(id=group_id)
    users = User.objects.filter(
        Q(stu_id=sid) | Q(tel=sid) | Q(id_card=sid) | Q(email=sid)
    )
    if users.exists():
        new_user = users.first()
        grp.user_set.add(new_user)
    return JR(success_msg('ok'))

def group_delete_user_by_data(request, group_id):
    data = json.loads(request.body)
    users_list = data
    print(users_list)
    grp = Group.objects.get(id=group_id)
    grp.user_set.clear()
    for user in users_list:
        user_obg = User.objects.get(stu_id=user['stu_id'])
        grp.user_set.add(user_obg)
    return JR(success_msg("ok"))

def get_info(request):
    ret_dict = generate_user_info(request)
    return JR(ret_dict)
