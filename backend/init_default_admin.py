# encoding=utf-8

__doc__ = """本脚本仅能在django已经连接数据库的情况下才能独立使用"""

# 脚本外使用django代码需要初始化django环境
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'exam_system.settings')
import django
django.setup()

import uuid
import random

from exam_system.settings import MEDIA_ROOT
from system_manage.models import SystemManageSetting

from openpyxl import load_workbook

# 执行django代码

# 创建默认用户（超级用户）
from user.models import User
from module.models import Module, SubModule, ModuleRole

def init_admin():
    password = str(uuid.uuid5(uuid.NAMESPACE_DNS, str(uuid.uuid1()) + str(random.random()))).replace('-','')
    default_admin_data = {
        'password': password,
        'name': 'admin',
        'stu_id': 'admin',
        'tel': '1234567',
        'id_card': '1234567',
        'email': 'admin@123.com',
        'role': 'admin',
    }
    # 若没有默认admin账户，则创建随机密码的admin账户
    if User.objects.filter(stu_id='admin').exists() == False:
        User.objects.create_superuser(
            password=default_admin_data['password'],
            name=default_admin_data['name'],
            stu_id=default_admin_data['stu_id'],
            tel=default_admin_data['tel'],
            id_card=default_admin_data['id_card'],
            email=default_admin_data['email'],
            role=default_admin_data['role'],
        )
        print("create super user successfully!")
        print("[!] default username & password: ", default_admin_data['stu_id'], default_admin_data['password'])
    else:
        # 存在admin默认账户
        print("[!] admin has existed!")

def init_menu():
    if len(Module.objects.all()) <= 2:
        roles_list_str = ""
        menu_excel_file_path = os.path.join(MEDIA_ROOT, 'template', 'EVO.xlsx')
        # menu_excel_file_path = os.path.join(r"C:\Users\ranja\Desktop\WorkTask\exam_system\backend\volume\media\template", 'EVO.xlsx')
        wb = load_workbook(menu_excel_file_path)
        handle_sheet = wb.worksheets[0]
        # 从第二行开始读取
        for row in handle_sheet.iter_rows(min_row=2, values_only=True):
            if not Module.objects.filter(menus=row[0], ).exists():
                # 避免插入重复 或 不存在的条目
                # 新建一级菜单
                new_module_obj = Module.objects.create(
                    menus=row[0],
                    description=row[1],
                    order=row[2],
                    icon=row[3],
                )
            first_module_obj = Module.objects.get(menus=row[0])
            if not SubModule.objects.filter(menus=first_module_obj, child=row[4], ).exists():
                new_sub_module_obj = SubModule.objects.create(
                    # 避免插入重复 或 不存在的条目
                    # 新建二级菜单
                    menus=first_module_obj,
                    child=row[4],
                    description=row[5],
                    frontpath=row[7],
                    icon=row[6],
                    order=row[8],
                    role_can_get=row[9],
                )
                if row[9] != None and row[9] != '':
                    roles_list_str = roles_list_str + "," + row[9]
        roles_list = roles_list_str.split(',')  # 使用逗号分割字符串并创建单词列表
        roles_set = set(role for role in roles_list if role not in (None, ''))  # 使用集合推导式去除 None 和 ''
        roles_set.add("admin")
        for role in roles_set:
            ModuleRole.objects.create(
                role=role,
            )
        print("[!] init menu successfully!")
    else:
        print("[!] menu has existed!")

def init_system_settings():
    if SystemManageSetting.objects.filter(id=1).exists() == False:
        sm_obj = SystemManageSetting.objects.create(
            id=1,
            main_title="EVO",
            vice_title="evo site created by senranja",
            login_title="evo login",
            cheat_mouse_out=10,
            cheat_page_out=10,
        )

        print("[!] init system_settings successfully!")
    else:
        print("[!] system_settings has existed!")

def main():
    # 【初始化admin账户】
    print("[!] init admin account!")

    try:
        init_admin()
    except Exception as e:
        print("create super user failed!", str(e))

    # 【初始化菜单】
    print("[!] init menu")
    try:
        init_menu()
    except Exception as e:
        print("init menu failed!", str(e))

    # 【初始化系统设置】
    print("[!] init system settings")
    try:
        init_system_settings()
    except Exception as e:
        print("init system settings failed!", str(e))

if __name__ == '__main__':
    main()