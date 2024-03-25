# encoding=utf-8

import json
import os
import time

from django.http import JsonResponse as JR
from conf.errors import error_log_msg, success_msg, error_excel_import_questions_msg

from openpyxl import load_workbook

# Create your views here.
from exam_manage.models import ExamManage, ExamManage_QuestionDatabase
from subject_manage.models import Subject
from question_manage.models import ExamQuestion
from exam_system.settings import EXCEL_IMPORT_QUESTION_ROOT

def handle_difficulty_level(content:str):
    content = content.strip().lower()
    easy_list = ['简单','容易','低', 'easy']
    middle_list = ['中等','中','middle',]
    hard_list = ['困难','高','难','high',]
    if content in easy_list:
        return 'easy'
    if content in middle_list:
        return 'middle'
    if content in hard_list:
        return 'hard'

def question_database_list(request):
    unique_question_databases = ExamQuestion.objects.values_list('question_database', flat=True).distinct()
    return JR({"data": [ str(i) for i in unique_question_databases]})

def update_exam_question(request, exam_id):
    """更新 考试--题库 对应关系。将 考试 的全部题库 对应全部删除，然后再设置入"""
    data = json.loads(request.body)
    # 获取 考试id 外键
    em_obj = ExamManage.objects.get(id=exam_id)
    # 删除全部旧有数据
    eqs_by_exam_id = ExamManage_QuestionDatabase.objects.filter(
        exam_manage=em_obj,
    )
    eqs_by_exam_id.delete()
    # 计算总分
    max_score = 0
    # 新增修改数据
    for item in data:
        if item['question_database']=="" or item['question_database']==None:
            pass
        else:
            ExamManage_QuestionDatabase.objects.create(
                exam_manage=em_obj,
                difficulty_level=item['difficulty_level'],
                question_database=item['question_database'],
                question_num=item['question_num'],
                question_type=item['question_type'],
                single_question_score=item['single_question_score'],
            )
            max_score = max_score + float(item['question_num']) * float(item['single_question_score'])
    # 更新考试满分
    em_obj.max_score = max_score
    em_obj.save()
    return JR(success_msg("更新成功"))

def import_question_from_excel(request):
    try:
        excel_file = request.FILES['questions_exam_file']

        user = request.user
        id = user.id
        user_name = user.name
        # print(excel_file)  # 打印文件名

        # 将文件保存到某位置
        current_time = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
        filename = "试题导入_{current_time}_{id}_{name}.xlsx".format(current_time=current_time,id=id,name=user_name)
        file_path = os.path.join(EXCEL_IMPORT_QUESTION_ROOT, filename)
        with open(file_path, 'wb') as f:
            for chunk in excel_file.chunks():
                f.write(chunk)

        # excel文件中内容数据入库
        wb = load_workbook(excel_file)
        handle_sheet = wb.worksheets[0]
        # 从第二行开始读取
        for row in handle_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]!="" and row[0]!=None and row[3]!="" and row[3]!=None:
                # 避免读取空值行
                if not Subject.objects.filter(subject=row[8]).exists():
                    Subject.objects.create(subject=row[8])
                subject = Subject.objects.get(subject=row[8])
                if not ExamQuestion.objects.filter(
                        question_type=row[0],
                        reference_answer=row[1],
                        question_name=row[2],
                        question_text=row[3],
                        choice_text=row[4],
                        question_image_data=row[5],
                        difficulty_level=handle_difficulty_level(row[6]),
                        description=row[7],
                        subject=subject,
                        question_database=row[9],
                        author=row[10],
                ).exists():
                    # 避免插入重复题目、题库
                    exam_question_obj = ExamQuestion.objects.create(
                        question_type=row[0],
                        reference_answer=row[1],
                        question_name=row[2],
                        question_text=row[3],
                        choice_text=row[4],
                        question_image_data=row[5],
                        difficulty_level=handle_difficulty_level(row[6]),
                        description=row[7],
                        subject=subject,
                        question_database=row[9],
                        author=row[10],
                    )
                    # for cell_value in row:
                    #     print(cell_value)

        # 手动关闭文件句柄
        excel_file.close()

        return JR(success_msg('import_question_from_excel successfully'))
    except Exception as err:
        return JR(error_excel_import_questions_msg(request, 'import_question_from_excel failed: '+str(err)))
