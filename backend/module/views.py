# encoding=utf-8
import json
import os
import time

from openpyxl import load_workbook

from django.contrib.auth.models import Group
from django.http import JsonResponse as JR
from conf.errors import success_msg
from guardian.shortcuts import assign_perm
from module.models import Module, SubModule, ModuleRole
from exam_system.settings import EXCEL_IMPORT_QUESTION_ROOT


def test(request):
    user = request.user
    ret_module_list = []
    pm = 'module.can_get'
    modules = Module.objects.all().order_by('order')
    for module in modules:
        if user.has_perm(pm, module):
            ret_module_list.append(module)
    print("授权列表", ret_module_list)
    ret = success_msg("ok")
    return JR(ret)

def get_roles_list(request):
    # 获取module_role中的全部角色权限
    ret = success_msg("ok")

    roles_list = []
    roles = ModuleRole.objects.all()
    for role in roles:
        roles_list.append(role.role)

    ret.update({"data": roles_list})
    return JR(ret)

def import_excel(request):
    excel_file = request.FILES['exam_file']

    user = request.user
    id = user.id
    user_name = user.name

    # module 角色列表
    roles_list_str = ""

    # 将文件保存到某位置
    current_time = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
    filename = "菜单模块_{current_time}_{id}_{name}.xlsx".format(current_time=current_time,id=id,name=user_name)
    file_path = os.path.join(EXCEL_IMPORT_QUESTION_ROOT, filename)
    with open(file_path, 'wb') as f:
        for chunk in excel_file.chunks():
            f.write(chunk)

    # excel文件中内容数据入库
    wb = load_workbook(excel_file)
    handle_sheet = wb.worksheets[0]
    # 从第二行开始读取
    for row in handle_sheet.iter_rows(min_row=2, values_only=True):
        if not Module.objects.filter(menus=row[0],).exists():
            # 避免插入重复 或 不存在的条目
            # 新建一级菜单
            new_module_obj = Module.objects.create(
                menus=row[0],
                description=row[1],
                order=row[2],
                icon=row[3],
            )
        first_module_obj = Module.objects.get(menus=row[0])
        if not SubModule.objects.filter(menus=first_module_obj, child=row[4],).exists():
            new_sub_module_obj = SubModule.objects.create(
                # 避免插入重复 或 不存在的条目
                # 新建二级菜单
                menus=first_module_obj,
                child=row[4],
                description=row[5],
                frontpath=row[7],
                icon=row[6],
                order=row[8],
                role_can_get=row[9],
            )
            if row[9]!=None and row[9]!='':
                roles_list_str = roles_list_str + "," + row[9]

    # 手动关闭文件句柄
    excel_file.close()

    roles_list = roles_list_str.split(',') # 使用逗号分割字符串并创建单词列表
    roles_set = set(role for role in roles_list if role not in (None, ''))  # 使用集合推导式去除 None 和 ''
    roles_set.add("admin")
    for role in roles_set:
        ModuleRole.objects.create(
            role=role,
        )

    return JR(success_msg('import menu excel'))

def generate_user_info(request):
    '''返回用户基本信息、menus、ruleNames'''
    user = request.user
    ret = success_msg("ok")
    ret.update(
        {
            "data": {
                "id": user.id,
                "username": user.name,
                "avatar": user.avatar,
                "super": user.is_superuser,
                "role": user.role,
            }
        }
    )
    menus = []
    for single_module_obj in Module.objects.all():
        # 一级菜单 循环 二级菜单

        # 二级菜单 sub_modules
        sub_modules_obj_list = []
        for single_sub_obj in SubModule.objects.filter(menus=single_module_obj):
            # 用户有二级菜单权限才可获取此二级菜单
            # 可访问条件：符合role角色，或者有 perm权限
            if single_sub_obj.role_can_get!=None:
                role_get = single_sub_obj.role_can_get.split(',')
            else:
                role_get = []

            # 去除 '' 和 None
            role_get_list = [item for item in role_get if item is not None and item != '']

            # pm = 'module.custom_submodule_can_get'
            # if user.has_perm(pm, single_sub_obj) or user.role in role_get_list
            if user.role in role_get_list or user.role=='admin':
                sub_modules_obj_list.append(
                    {
                        "name": single_sub_obj.child,
                        "desc": single_sub_obj.description,
                        "frontpath": single_sub_obj.frontpath,
                        "order": single_sub_obj.order,
                        "icon": single_sub_obj.icon,
                    }
                )
        if sub_modules_obj_list != []:
            # 保证不出现 一级菜单存在但是子菜单为[] 的情况。二级菜单不为空，才添加。

            # 二级菜单根据order排序
            sub_modules_obj_list = sorted(sub_modules_obj_list, key=lambda x: x["order"])
            # 一级菜单
            menus.append(
                {
                    "name": single_module_obj.menus,
                    "description": single_module_obj.description,
                    "icon": single_module_obj.icon,
                    "order": single_module_obj.order,
                    "frontpath": None,
                    "child": sub_modules_obj_list,
                }
            )
        elif sub_modules_obj_list == []:
            pass
    # 一级菜单根据order排序
    menus = sorted(menus, key=lambda x: x["order"])

    ruleNames = [
    ]

    ret["data"].update({'menus': menus})
    ret["data"].update({'ruleNames': ruleNames})

    return ret

